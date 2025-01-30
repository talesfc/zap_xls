import os
import sys
import importlib.util
import datetime
from pathlib import Path

import re

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Alignment
from openpyxl.styles.fonts import _no_value

import Chat_Utils as UTILS_CHAT

# Prepara planilha de resultado

def build_WBook(templatefile):
    wb = load_workbook(filename = templatefile)
    return(wb)

def saveXLSWBook(wbOUT, outfile):
    wbOUT.save(filename=outfile)    
    return()

def insert_notesheet(wsheet, i, reference, note):
    datahora= datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    i=i+1
    wsheet.cell(row= i, column= 1, value= i)
    wsheet.cell(row= i, column= 2, value= datahora)
    wsheet.cell(row= i, column= 3, value= reference)
    wsheet.cell(row= i, column= 4, value= note)
    return(wsheet)

def insert_eventsheet(wsheet, i, agent, evento, datahora):
    i=i+1
    wsheet.cell(row= i, column= 1, value= i)
    wsheet.cell(row= i, column= 2, value= datahora)
    wsheet.cell(row= i, column= 3, value= agent)
    wsheet.cell(row= i, column= 4, value= evento)
    # atext= aline.strip()
    # for tipo in UTILS_CHAT.event_list:
    #     pos= atext.find(tipo)
    #     if pos > 0:
    #         nome= atext[:pos]
    #         evento= atext[pos:]
    #         wsheet.cell(row= i, column= 3, value= nome)
    return(wsheet)

def insert_postsheet(wsheet, i, aline, datahora):
    i=i+1
    n= len(datahora)
    if n > 10:
        text= aline[n:]
        marca= text.find(":")
        poster= text[:marca]
        message= text[marca+1:]
        wsheet.cell(row= i, column= 1, value= i)
        wsheet.cell(row= i, column= 2, value= datahora)
        wsheet.cell(row= i, column= 5, value= message)
        # se o poster contém apenas números é phone, se não é name
        phone= poster.replace("+","").replace("-","").replace(" ","").replace("(","").replace(")","")
        if re.match("[0-9]+$", phone):
            #print(f"'{phone}' contém apenas números.")
            wsheet.cell(row= i, column= 3, value= phone)
        else:
            #print(f"'{poster}' não contém apenas números.")
            name= poster.replace("- ","")
            wsheet.cell(row= i, column= 4, value= name)
    else:
        wsheet.cell(row= i, column= 1, value= i)
        wsheet.cell(row= i, column= 5, value= aline)
    return(wsheet)






